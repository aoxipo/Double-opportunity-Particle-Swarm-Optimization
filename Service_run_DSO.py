# -*- coding: utf-8 -*-
# @Author: Administrator
# @Date:   2019-12-31 09:52:17
# @Last Modified by:   Administrator
# @Last Modified time: 2020-01-02 21:07:48
# 
import pickle
import logging
from datetime import datetime
from config import *
from Service_Algorithm import *
from openpyxl import Workbook
from multiprocessing import Pool
from data_process.generate_topology import generate_topology
from utils import *

'''
BBO设定
将上海市金茂大厦作为最核心地带，其经纬度为121.50109,31.23691。
根据欧式距离计算，当距金茂大厦20千米以内，
地租成本约为500，每隔20km，
成本线性降低100；
服务器成本设置为100.

'''


def run(edge, dis):
    print("start run")
    
    bbo = BBODeployer(edge_list=edge, distances=dis,particle_size=InitParticle);  #particle_size=20
    print("变覆盖范围start");
    filepath_coverage=filepath+"coverage_bbo"
    if(set_coverage):
        with open(filepath_coverage+".txt", 'w') as f:  # 变覆盖范围
            wb = Workbook()
            ws = wb.active
            ws.append(['service_MAX_WORKLOAD', 'SERVICE_COVERAGE', 'MAX_TOTAL_PRICE', 'MAX_WORKOLAD_BALANCE', None,
                       'bbo.average_delay', 'bbo.total_price', 'bbo.wokload_balance','bbo.resource_utilization','bbo.srv_number','spfa jump'])
    
            for coverage in range(*global_coverage):
                for n in range(global_smallnumber_base,global_number_base,200):
                    bbo.set_parameter(service_coverage=coverage,base_station_num=n)
                    placed_services = bbo.service_deployer().get_all_services()
        
                    # txt
                    bbo.print_result(result_services=placed_services)
                    bbo.print_result(file=f, result_services=placed_services)
        
                    # csv
                    csv_line = bbo.csv_data()
                    # 如果5000次未找到解，返回的是空list
                    if not placed_services:
                        csv_line[-3:] = 0,0,0
                    ws.append(csv_line)
             
            wb.save(filepath_coverage+".xlsx")

    # 变 最高价格
    if(set_price):
        print("最高价格start");
        filepath_price=filepath+"price_bbo"
        with open(filepath_price+".txt", 'w') as f:
            
            wb = Workbook()
            ws = wb.active
            ws.append(['service_MAX_WORKLOAD', 'SERVICE_COVERAGE', 'MAX_TOTAL_PRICE', 'MAX_WORKOLAD_BALANCE', None,
                       'bbo.average_delay', 'bbo.total_price', 'bbo.wokload_balance','bbo.srv_number','bbo.resource_utilization'])
    
            for price in range(*global_price):
                bbo.set_parameter(price=price)
                placed_services = bbo.service_deployer().get_all_services()
                # txt
                bbo.print_result(result_services=placed_services)
                bbo.print_result(file=f, result_services=placed_services)
    
                # csv
                csv_line = bbo.csv_data()
                # 如果5000次未找到解，返回的是空list
                if not placed_services:
                    csv_line[-3:] = 0, 0, 0
                ws.append(csv_line)
            
            wb.save(filepath_price+".xlsx")

    # 变 最大负载均衡
    if(set_balance):
        print("变最大负载均衡start");
        filepath_balance=filepath+"workload_balance_bbo"
        with open(filepath_balance+'.txt', 'w') as f:
    
            wb = Workbook()
            ws = wb.active
            ws.append(['service_MAX_WORKLOAD', 'SERVICE_COVERAGE', 'MAX_TOTAL_PRICE', 'MAX_WORKOLAD_BALANCE', None,
                       'bbo.average_delay', 'bbo.total_price', 'bbo.wokload_balance','bbo.srv_number','bbo.resource_utilization'])
    
            for workload_balance in range(*global_loadbalance):
                bbo.set_parameter(workload_balance=workload_balance)
                placed_services = bbo.service_deployer().get_all_services()
    
                # txt
                bbo.print_result(result_services=placed_services)
                bbo.print_result(file=f, result_services=placed_services)
    
                # csv
                csv_line = bbo.csv_data()
                # 如果5000次未找到解，返回的是空list
                if not placed_services:
                    csv_line[-3:] = 0, 0, 0
                ws.append(csv_line)
    
            wb.save(filepath_balance+'.xlsx')

    # 变最大负载
    if(set_workload):
        print("变最大负载start");
        filepath_maxworkload=filepath+"max_workload_bbo"
        with open(filepath_maxworkload+'.txt', 'w') as f:
    
            wb = Workbook()
            ws = wb.active
            ws.append(['service_MAX_WORKLOAD', 'SERVICE_COVERAGE', 'MAX_TOTAL_PRICE', 'MAX_WORKOLAD_BALANCE', None,
                       'bbo.average_delay', 'bbo.total_price', 'bbo.wokload_balance','bbo.srv_number','bbo.resource_utilization'])
    
            for max_workload in global_workload:
                bbo.set_parameter(max_workload=max_workload)
                placed_services = bbo.service_deployer().get_all_services()
    
                # txt
                bbo.print_result(result_services=placed_services)
                bbo.print_result(file=f, result_services=placed_services)
    
                # csv
                csv_line = bbo.csv_data()
                # 如果5000次未找到解，返回的是空list
                if not placed_services:
                    csv_line[-3:] = 0, 0, 0
                ws.append(csv_line)
    
            wb.save(filepath_maxworkload+'.xlsx')



if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO)

    start_time = "start run at: {0}".format(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

    print('开始service deploym，topology---->max_link:{}  max_dis:{}'.format(max_link, max_dis))
#对比 基站数目更改 100-1300 200
#改变最大时延从2km到9km，

    # 载入edge信息
    # 格式：list每个元素是edge_server类
    edge_info_name = '../cache/edge_info_l' + str(max_link) + '_d' + str(max_dis)
    edge_info_name = '../cache/edge_info'
    with open(edge_info_name, 'rb') as f:
        edge_list = pickle.load(f)

    # 载入基站信息
    # list每个元素是base_station类
    with open(r'../cache/base_stations_with_user_info', 'rb') as f:
        bs_list = pickle.load(f)['value']
    for obj in bs_list:
        obj.base_station_id=obj.id
    # 载入基站距离矩阵
    # dis_list[20][1] = print(dis_list['value'][1][20]) =基站1和基站20距离

    # with open(r'../cache/topology', 'rb') as f:
    #     dis_list = pickle.load(f)['value']
    #
    data = DataUtils('../data/基站经纬度_修改完整版_增加缺失地址_修改重复地址.csv', '../data/上网信息输出表（日表）7月15号之后.csv')
    topology = generate_topology(data)
    run(bs_list, topology)

    end_time = "end run at:{0}".format(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

    print('\n\n\n')
    logging.info(start_time)
    logging.info(end_time)
